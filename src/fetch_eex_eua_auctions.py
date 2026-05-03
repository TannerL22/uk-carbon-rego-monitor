"""Fetch and standardise official EEX EUA primary auction results.

The EEX public auction workbooks are treated as raw public source files. This
script parses successful EUA primary auctions into the dashboard's standard
carbon-auction schema and replaces the output CSV only after a successful parse.
"""

from __future__ import annotations

import csv
from collections import defaultdict
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CARBON_RAW = ROOT / "data" / "raw" / "carbon"
SOURCE_REGISTER = ROOT / "data" / "source_register.csv"
OUTPUT = CARBON_RAW / "eua_auction_results.csv"

BASE_URL = "https://public.eex-group.com/eex/eua-auction-report"
FILE_PATTERN = "emission-spot-primary-market-auction-report-{year}-data.xlsx"
TIMEOUT_SECONDS = 45

OUTPUT_COLUMNS = [
    "market",
    "auction_date",
    "auction_volume",
    "clearing_price",
    "currency",
    "cover_ratio",
    "reference_price",
    "source",
    "source_url",
    "downloaded_at",
    "notes",
]


class EexFetchError(RuntimeError):
    """Raised when official EEX auction data cannot be fetched or parsed."""


def iso_z(value: datetime) -> str:
    return value.replace(second=0, microsecond=0).strftime("%Y-%m-%dT%H:%MZ")


def report_url(year: int) -> str:
    return f"{BASE_URL}/{FILE_PATTERN.format(year=year)}"


def fetch_workbook(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": "uk-carbon-rego-monitor/1.0"})
    try:
        with urlopen(request, timeout=TIMEOUT_SECONDS) as response:
            return response.read()
    except (HTTPError, URLError, TimeoutError) as exc:
        raise EexFetchError(f"EEX auction report request failed for {url}: {exc}") from exc


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    parsed = datetime.strptime(str(value).strip()[:10], "%Y-%m-%d")
    return parsed.date()


def as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows):
        headers = {normalize_header(value): position for position, value in enumerate(row)}
        if "date" in headers and "auction_price_€/tco2" in headers and "auction_volume_tco2" in headers:
            return index, headers
    raise EexFetchError("EEX workbook did not include the expected primary-auction header row.")


def parse_workbook(payload: bytes, url: str) -> list[dict[str, object]]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    if "Primary Market Auction" not in workbook.sheetnames:
        raise EexFetchError("EEX workbook did not include a 'Primary Market Auction' worksheet.")

    worksheet = workbook["Primary Market Auction"]
    rows = list(worksheet.iter_rows(values_only=True))
    header_row, headers = find_header_row(rows)

    price_col = headers["auction_price_€/tco2"]
    volume_col = headers["auction_volume_tco2"]
    cover_col = headers.get("cover_ratio")
    status_col = headers.get("status")
    country_col = headers.get("country")

    parsed: list[dict[str, object]] = []
    for row in rows[header_row + 1 :]:
        auction_date = parse_date(row[headers["date"]])
        price = as_float(row[price_col])
        volume = as_float(row[volume_col])
        if auction_date is None or price is None or volume is None:
            continue

        status = str(row[status_col] if status_col is not None else "successful").strip().lower()
        if status and status != "successful":
            continue

        country = str(row[country_col] if country_col is not None else "").strip()
        parsed.append(
            {
                "auction_date": auction_date.isoformat(),
                "auction_volume": volume,
                "clearing_price": price,
                "cover_ratio": as_float(row[cover_col]) if cover_col is not None else None,
                "country": country,
                "source_url": url,
            }
        )

    if not parsed:
        raise EexFetchError(f"EEX workbook {url} did not yield any successful EUA auction rows.")
    return parsed


def aggregate_daily(rows: list[dict[str, object]], downloaded_at: str) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["auction_date"])].append(row)

    output: list[dict[str, object]] = []
    for auction_date in sorted(grouped):
        items = grouped[auction_date]
        total_volume = sum(float(item["auction_volume"]) for item in items)
        weighted_price = sum(float(item["clearing_price"]) * float(item["auction_volume"]) for item in items) / total_volume
        cover_values = [float(item["cover_ratio"]) for item in items if item.get("cover_ratio") is not None]
        weighted_cover = (
            sum(float(item["cover_ratio"]) * float(item["auction_volume"]) for item in items if item.get("cover_ratio") is not None)
            / sum(float(item["auction_volume"]) for item in items if item.get("cover_ratio") is not None)
            if cover_values
            else ""
        )
        source_urls = sorted({str(item["source_url"]) for item in items})
        countries = sorted({str(item["country"]) for item in items if str(item.get("country", "")).strip()})
        output.append(
            {
                "market": "EUA",
                "auction_date": auction_date,
                "auction_volume": int(round(total_volume)),
                "clearing_price": round(weighted_price, 2),
                "currency": "EUR",
                "cover_ratio": round(weighted_cover, 2) if weighted_cover != "" else "",
                "reference_price": "",
                "source": "EEX EUA Primary Market Auction Report",
                "source_url": "; ".join(source_urls),
                "downloaded_at": downloaded_at,
                "notes": (
                    "Official EEX public workbook parsed by build script; successful EUA primary auctions "
                    f"aggregated by auction date across {', '.join(countries) if countries else 'listed auction regions'}."
                ),
            }
        )
    return output


def write_csv_atomically(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", newline="", encoding="utf-8", delete=False, dir=path.parent, suffix=".tmp") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
        temp_path = Path(handle.name)
    temp_path.replace(path)


def update_source_register(rows: list[dict[str, object]], downloaded_at: str) -> None:
    if not SOURCE_REGISTER.exists():
        return

    with SOURCE_REGISTER.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        register_rows = list(reader)
        fieldnames = reader.fieldnames or []

    if not fieldnames:
        return

    start = str(rows[0]["auction_date"])
    end = str(rows[-1]["auction_date"])
    row_found = False
    for row in register_rows:
        if row.get("source_id") == "SRC-004":
            row_found = True
            row["dataset_name"] = "EEX EUA auction results"
            row["source_owner"] = "EEX"
            row["source_type"] = "official public auction report"
            row["source_url"] = BASE_URL
            row["manual_or_api"] = "public file"
            row["downloaded_at"] = downloaded_at
            row["published_at"] = end
            row["data_period_start"] = start
            row["data_period_end"] = end
            row["used_for"] = "EUA primary auction signal"
            row["known_limitations"] = "Official EEX primary auction report parsed from public Excel workbooks; not a live secondary-market feed."

    if not row_found:
        register_rows.append(
            {
                "source_id": "SRC-004",
                "dataset_name": "EEX EUA auction results",
                "source_owner": "EEX",
                "source_type": "official public auction report",
                "source_url": BASE_URL,
                "manual_or_api": "public file",
                "downloaded_at": downloaded_at,
                "published_at": end,
                "data_period_start": start,
                "data_period_end": end,
                "used_for": "EUA primary auction signal",
                "known_limitations": "Official EEX primary auction report parsed from public Excel workbooks; not a live secondary-market feed.",
            }
        )

    with SOURCE_REGISTER.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(register_rows)


def main() -> None:
    current_year = datetime.now(UTC).year
    years = [current_year - 2, current_year - 1, current_year]
    all_rows: list[dict[str, object]] = []
    errors: list[str] = []

    for year in years:
        url = report_url(year)
        try:
            all_rows.extend(parse_workbook(fetch_workbook(url), url))
        except EexFetchError as exc:
            errors.append(str(exc))

    if errors:
        raise EexFetchError("One or more EEX EUA auction reports could not be fetched. Existing CSV was left unchanged. " + " | ".join(errors))

    if not all_rows:
        raise EexFetchError("No EEX EUA auction rows could be fetched. Existing CSV was left unchanged. " + " | ".join(errors))

    downloaded_at = iso_z(datetime.now(UTC))
    rows = aggregate_daily(all_rows, downloaded_at)
    write_csv_atomically(OUTPUT, rows)
    update_source_register(rows, downloaded_at)
    print(f"Fetched {len(rows)} EEX EUA auction dates from {rows[0]['auction_date']} to {rows[-1]['auction_date']}.")


if __name__ == "__main__":
    main()
